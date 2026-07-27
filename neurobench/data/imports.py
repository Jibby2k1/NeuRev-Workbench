"""Safe, metadata-first dataset import and local-registration contracts.

This module deliberately stops before scientific processing.  It records the
source, checksum, observed file metadata, unresolved scientific fields, and a
truthful lifecycle state so the browser can offer the next safe action without
inventing modality or acquisition facts.
"""
from __future__ import annotations

import csv
from datetime import datetime, timezone
import hashlib
import json
import os
from pathlib import Path
import re
from typing import Any, Iterable, Mapping
import uuid

from neurobench.data.video import video_metadata
from neurobench.validation.schemas import validate_dict


IMPORT_SCHEMA_VERSION = 1
MAX_IMPORT_BYTES = 20_000_000_000
MAX_IMPORT_RECORD_BYTES = 8 * 1024 * 1024
MAX_LABEL_ROWS = 1_000_000
MAX_LABEL_ARTIFACT_BYTES = 100_000_000
MAX_NEUREV_JSON_BYTES = 64 * 1024 * 1024
SUPPORTED_VIDEO_SUFFIXES = frozenset({".npy", ".tif", ".tiff"})
SUPPORTED_LABEL_SUFFIXES = frozenset({".csv", ".tsv", ".xlsx"})
SUPPORTED_NEUREV_SUFFIXES = frozenset({".json"})
IMPORT_STATES = (
    "uploaded",
    "metadata_needed",
    "qc_ready",
    "processing",
    "ready",
    "annotation_in_progress",
    "complete",
    "failed",
)
IMPORT_STATE_TRANSITIONS = {
    "uploaded": frozenset({"metadata_needed", "qc_ready", "failed"}),
    "metadata_needed": frozenset({"metadata_needed", "qc_ready", "failed"}),
    "qc_ready": frozenset({"qc_ready", "processing", "failed"}),
    "processing": frozenset({"ready", "complete", "failed"}),
    "ready": frozenset({"ready", "annotation_in_progress", "complete", "failed"}),
    "annotation_in_progress": frozenset({"annotation_in_progress", "complete", "failed"}),
    "complete": frozenset({"complete"}),
    "failed": frozenset({"failed", "metadata_needed", "qc_ready"}),
}
DATASET_ID_RE = re.compile(r"[^a-z0-9_-]+")


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="microseconds").replace("+00:00", "Z")


def normalize_dataset_id(value: str, *, fallback: str = "dataset") -> str:
    """Return a stable filesystem-safe dataset identifier."""

    raw = str(value or "").strip().lower().replace(" ", "-")
    cleaned = DATASET_ID_RE.sub("-", raw).strip("-_.")
    return cleaned[:120] or normalize_dataset_id(fallback, fallback="dataset")


def import_id() -> str:
    """Return a collision-resistant import identifier."""

    return f"imp_{uuid.uuid4().hex[:16]}"


def _import_record_object_without_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"Import sidecar contains a duplicate object key: {key}")
        result[key] = value
    return result


def read_import_record(
    path: str | Path,
    *,
    expected_dataset_id: str,
    expected_app_dir: str | Path,
    workspace_root: str | Path,
    max_bytes: int = MAX_IMPORT_RECORD_BYTES,
) -> dict[str, Any]:
    """Read one sidecar and enforce its schema and storage identity.

    The filename is part of the durable identity: ``<import_id>.json``.  The
    declared dataset and app directory must also match the caller's resolved
    route/catalog context.  Callers that enumerate records may catch
    ``ValueError``/``OSError`` and omit invalid sidecars, thereby failing
    closed without treating corrupt local files as usable imports.
    """

    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Import sidecar does not exist: {source}")
    if source.stat().st_size > max_bytes:
        raise ValueError(f"Import sidecar exceeds the {max_bytes:,}-byte safety limit.")
    try:
        with source.open("rb") as handle:
            raw = handle.read(max_bytes + 1)
        if len(raw) > max_bytes:
            raise ValueError(f"Import sidecar exceeds the {max_bytes:,}-byte safety limit.")
        payload = json.loads(
            raw.decode("utf-8-sig"),
            object_pairs_hook=_import_record_object_without_duplicate_keys,
        )
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"Import sidecar is not valid JSON: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError("Import sidecar must contain one JSON object.")
    try:
        validate_dict(payload, "dataset_import")
    except Exception as exc:
        raise ValueError(f"Import sidecar failed schema validation: {exc}") from exc

    record = dict(payload)
    import_id_value = str(record.get("import_id") or "")
    if source.name != f"{import_id_value}.json":
        raise ValueError(
            f"Import sidecar filename/import_id mismatch: {source.name} != {import_id_value}.json"
        )
    dataset_id = str(record.get("dataset_id") or "")
    if dataset_id != str(expected_dataset_id):
        raise ValueError(
            f"Import sidecar dataset_id mismatch: {dataset_id or '(missing)'} != {expected_dataset_id}"
        )

    declared_app = str(record.get("app_dir") or "")
    if not declared_app:
        raise ValueError("Import sidecar app_dir is required for storage identity validation.")
    workspace = Path(workspace_root).expanduser().resolve()
    declared_path = Path(declared_app).expanduser()
    if not declared_path.is_absolute():
        declared_path = workspace / declared_path
    if declared_path.resolve() != Path(expected_app_dir).expanduser().resolve():
        raise ValueError(
            f"Import sidecar app_dir mismatch: {declared_path.resolve()} != "
            f"{Path(expected_app_dir).expanduser().resolve()}"
        )
    return record


def workspace_roots(workspace_root: str | Path) -> tuple[Path, ...]:
    """Return roots in which user data may be registered or created."""

    workspace = Path(workspace_root).expanduser().resolve()
    return tuple((workspace / name).resolve() for name in ("Inputs", "Outputs"))


def resolve_allowed_local_path(
    value: str | Path,
    *,
    workspace_root: str | Path,
    allowed_roots: Iterable[str | Path] | None = None,
) -> Path:
    """Resolve an existing local file and reject paths outside configured roots."""

    raw_candidate = Path(value).expanduser()
    if not raw_candidate.is_absolute():
        raw_candidate = Path(workspace_root).expanduser() / raw_candidate
    candidate = raw_candidate.resolve(strict=False)
    if not candidate.is_file():
        raise FileNotFoundError(f"Local source file does not exist: {candidate}")
    roots = tuple(Path(root).expanduser().resolve() for root in (allowed_roots or workspace_roots(workspace_root)))
    if not any(candidate == root or root in candidate.parents for root in roots):
        allowed = ", ".join(str(root) for root in roots)
        raise ValueError(f"Local source must be inside an allowed root ({allowed}); got {candidate}")
    return candidate


def relative_workspace_path(path: str | Path, *, workspace_root: str | Path) -> str:
    """Render a path relative to the workspace when possible."""

    resolved = Path(path).expanduser().resolve()
    try:
        return resolved.relative_to(Path(workspace_root).expanduser().resolve()).as_posix()
    except ValueError:
        return str(resolved)


def source_kind(path: str | Path) -> str:
    suffix = Path(path).suffix.lower()
    if suffix in SUPPORTED_VIDEO_SUFFIXES:
        return "video"
    if suffix in SUPPORTED_LABEL_SUFFIXES:
        return "label_table"
    if suffix in SUPPORTED_NEUREV_SUFFIXES:
        return "neurev_json"
    return "unsupported"


def _json_object_without_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"NeuRev JSON contains a duplicate object key: {key}")
        result[key] = value
    return result


def load_neurev_json(path: str | Path, *, max_bytes: int = MAX_NEUREV_JSON_BYTES) -> dict[str, Any]:
    """Load one bounded UTF-8 NeuRev JSON object without accepting duplicate keys."""

    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"NeuRev JSON source does not exist: {source}")
    size = source.stat().st_size
    if size > max_bytes:
        raise ValueError(f"NeuRev JSON exceeds the {max_bytes:,}-byte safety limit.")
    with source.open("rb") as handle:
        raw = handle.read(max_bytes + 1)
    if len(raw) > max_bytes:
        raise ValueError(f"NeuRev JSON exceeds the {max_bytes:,}-byte safety limit.")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError(f"NeuRev JSON must be valid UTF-8: {exc}") from exc
    try:
        payload = json.loads(text, object_pairs_hook=_json_object_without_duplicate_keys)
    except (json.JSONDecodeError, RecursionError) as exc:
        raise ValueError(f"Invalid NeuRev JSON: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError("NeuRev JSON must contain one top-level object.")
    return payload


def inspect_neurev_json(path: str | Path) -> dict[str, Any]:
    """Recognize and summarize one established NeuRev JSON contract.

    The adapter deliberately accepts only native review data, annotations v3,
    architecture-run manifests, and export-bundle manifests.  It never follows
    paths declared inside a payload and never returns the full payload as import
    metadata.
    """

    payload = load_neurev_json(path)
    keys = set(payload)
    signatures = [
        ("review_data", {"video", "parameters", "rois"}, "review_data"),
        ("annotations", {"schema_version", "rois", "events", "suggestions", "settings"}, "annotations"),
        ("architecture_runs", {"schema_version", "dataset_id", "runs"}, "architecture_runs"),
        (
            "export_bundle",
            {
                "schema_version",
                "export_bundle_id",
                "dataset_id",
                "run_ids",
                "created_at",
                "selection_policy",
                "alignment_status",
                "files",
                "provenance",
            },
            "export_bundle",
        ),
    ]
    matches = [(kind, schema_name) for kind, required, schema_name in signatures if required <= keys]
    if not matches:
        raise ValueError(
            "JSON is not a recognized NeuRev review_data, annotations v3, "
            "architecture_runs, or export_bundle object."
        )
    if len(matches) != 1:
        names = ", ".join(kind for kind, _ in matches)
        raise ValueError(f"NeuRev JSON mixes multiple native document shapes: {names}.")
    payload_kind, schema_name = matches[0]
    try:
        validate_dict(payload, schema_name)
    except Exception as exc:
        raise ValueError(f"NeuRev {payload_kind} JSON failed schema validation: {exc}") from exc

    declared_dataset_id: str | None = None
    counts: dict[str, int] = {}
    if payload_kind == "review_data":
        dataset = payload.get("dataset") if isinstance(payload.get("dataset"), Mapping) else {}
        parameters = payload.get("parameters") if isinstance(payload.get("parameters"), Mapping) else {}
        declared_dataset_id = str(
            dataset.get("dataset_id") or payload.get("dataset_id") or parameters.get("datasetId") or ""
        ) or None
        rois = payload.get("rois") if isinstance(payload.get("rois"), list) else []
        discovery = payload.get("discovery") if isinstance(payload.get("discovery"), Mapping) else {}
        suggestions = discovery.get("suggestions") if isinstance(discovery.get("suggestions"), list) else []
        counts = {
            "roi_count": len(rois),
            "event_count": sum(
                len(roi.get("events"))
                for roi in rois
                if isinstance(roi, Mapping) and isinstance(roi.get("events"), list)
            ),
            "suggestion_count": len(suggestions),
        }
    elif payload_kind == "annotations":
        for source_key, count_key in (
            ("rois", "roi_annotation_count"),
            ("events", "event_annotation_count"),
            ("suggestions", "suggestion_annotation_count"),
            ("virtualRois", "virtual_roi_count"),
            ("runs", "run_bucket_count"),
        ):
            value = payload.get(source_key)
            counts[count_key] = len(value) if isinstance(value, Mapping) else 0
    elif payload_kind == "architecture_runs":
        declared_dataset_id = str(payload.get("dataset_id") or "") or None
        counts = {"run_count": len(payload.get("runs") or [])}
    else:
        declared_dataset_id = str(payload.get("dataset_id") or "") or None
        counts = {
            "run_id_count": len(payload.get("run_ids") or []),
            "file_count": len(payload.get("files") or []),
        }
    return {
        "payload_kind": payload_kind,
        "payload_schema_version": payload.get("schema_version"),
        "declared_dataset_id": declared_dataset_id,
        "counts": counts,
    }


def checksum_file(path: str | Path, *, chunk_size: int = 8 * 1024 * 1024) -> dict[str, Any]:
    """Hash a source without loading it into memory."""

    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Source file does not exist: {source}")
    digest = hashlib.sha256()
    size = 0
    with source.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            size += len(chunk)
            digest.update(chunk)
    return {"sha256": digest.hexdigest(), "size_bytes": size}


def inspect_source(path: str | Path, *, workspace_root: str | Path | None = None) -> dict[str, Any]:
    """Inspect a source using bounded metadata readers and no scientific guesses."""

    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Source file does not exist: {source}")
    kind = source_kind(source)
    if kind == "neurev_json" and source.stat().st_size > MAX_NEUREV_JSON_BYTES:
        raise ValueError(f"NeuRev JSON exceeds the {MAX_NEUREV_JSON_BYTES:,}-byte safety limit.")
    checksum = checksum_file(source)
    observed: dict[str, Any] = {
        "format": source.suffix.lower().lstrip("."),
        "kind": kind,
        "original_name": source.name,
        "size_bytes": checksum["size_bytes"],
        "sha256": checksum["sha256"],
        "frame_rate_hz": None,
        "pixel_size_microns": None,
        "modality": None,
        "indicator": None,
    }
    warnings: list[str] = []
    if kind == "video":
        metadata = video_metadata(source)
        observed.update(
            {
                "shape": metadata.get("shape"),
                "frames": metadata.get("frames"),
                "height": metadata.get("height"),
                "width": metadata.get("width"),
                "dtype": metadata.get("dtype"),
                "storage_mode": metadata.get("storage_mode"),
            }
        )
        warnings.append("frame_rate_hz is unknown until supplied or read from trusted metadata")
        warnings.append("pixel_size_microns is unknown until supplied or read from trusted metadata")
    elif kind == "label_table":
        observed.update(inspect_label_table(source))
    elif kind == "neurev_json":
        observed.update(inspect_neurev_json(source))
        warnings.append("NeuRev JSON remains external until its bounded preview is explicitly confirmed")
        warnings.append("Declared paths inside NeuRev JSON are recorded only; they are never followed during import")
    else:
        raise ValueError(f"Unsupported source suffix: {source.suffix or '(none)'}")
    if workspace_root is not None:
        observed["source_path"] = relative_workspace_path(source, workspace_root=workspace_root)
    else:
        observed["source_path"] = str(source)
    return {"metadata": observed, "warnings": warnings}


def inspect_label_table(path: str | Path) -> dict[str, Any]:
    """Inspect label columns without discarding or normalizing rows."""

    source = Path(path).expanduser().resolve()
    suffix = source.suffix.lower()
    if suffix == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ModuleNotFoundError as exc:
            raise RuntimeError("XLSX label import requires openpyxl in the active environment.") from exc
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header = [str(value or "").strip() for value in next(rows, ())]
            count = 0
            for _ in rows:
                count += 1
                if count > MAX_LABEL_ROWS:
                    raise ValueError(f"Label table exceeds the {MAX_LABEL_ROWS:,}-row safety limit.")
        finally:
            workbook.close()
    else:
        delimiter = "\t" if suffix == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            header = [str(value or "").strip() for value in next(reader, [])]
            count = 0
            for _ in reader:
                count += 1
                if count > MAX_LABEL_ROWS:
                    raise ValueError(f"Label table exceeds the {MAX_LABEL_ROWS:,}-row safety limit.")
    if not header:
        raise ValueError("Label table has no header row.")
    return {
        "columns": header,
        "row_count": count,
        "label_mapping": infer_label_mapping(header),
    }


def infer_label_mapping(columns: Iterable[str]) -> dict[str, str | None]:
    """Suggest a non-destructive mapping for common label-table columns."""

    normalized = {str(column).strip().lower().replace(" ", "_"): str(column) for column in columns}
    aliases = {
        "roi_identity": ("roi_identity", "roi_id", "roi", "id", "cell_id", "neuron_id"),
        "roi_id": ("roi_identity", "roi_id", "roi", "id", "cell_id", "neuron_id"),
        "x": ("x", "centroid_x", "center_x", "x_px"),
        "y": ("y", "centroid_y", "center_y", "y_px"),
        "start_frame_ui": ("start_frame_ui", "start_frame", "start", "frame_start", "onset_frame"),
        "start_frame": ("start_frame_ui", "start_frame", "start", "frame_start", "onset_frame"),
        "end_frame_ui": ("end_frame_ui", "end_frame", "end", "frame_end", "offset_frame"),
        "end_frame": ("end_frame_ui", "end_frame", "end", "frame_end", "offset_frame"),
        "label": ("label", "state", "class", "annotation", "cell_state"),
        "confidence": ("confidence", "score", "probability"),
    }
    return {
        target: next((normalized[alias] for alias in choices if alias in normalized), None)
        for target, choices in aliases.items()
    }


def iter_label_rows(path: str | Path, *, limit: int = MAX_LABEL_ROWS) -> Iterable[dict[str, Any]]:
    """Yield label rows without loading a complete table into memory."""
    source = Path(path).expanduser().resolve()
    suffix = source.suffix.lower()
    if suffix != ".xlsx":
        delimiter = "\t" if suffix == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            for index, row in enumerate(reader):
                if index >= limit:
                    raise ValueError(f"Label table exceeds the {limit:,}-row safety limit.")
                yield {str(key).strip(): value for key, value in row.items() if key is not None}
        return
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("XLSX label import requires openpyxl in the active environment.") from exc
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        for index, values in enumerate(rows):
            if index >= limit:
                raise ValueError(f"Label table exceeds the {limit:,}-row safety limit.")
            yield {header: value for header, value in zip(headers, values) if header}
    finally:
        workbook.close()


def initial_state(metadata: Mapping[str, Any], *, qc_present: bool = False) -> str:
    """Choose a truthful lifecycle state from observed metadata only."""

    if metadata.get("kind") == "unsupported":
        return "failed"
    if metadata.get("kind") == "video" and (metadata.get("frames") and metadata.get("width") and metadata.get("height")):
        return "qc_ready" if qc_present else "metadata_needed"
    if metadata.get("kind") in {"label_table", "neurev_json"}:
        return "qc_ready"
    return "metadata_needed"


def make_import_record(
    *,
    dataset_id: str,
    import_id_value: str,
    source_mode: str,
    original_name: str,
    source_path: str,
    destination_path: str,
    metadata: Mapping[str, Any],
    warnings: Iterable[str] = (),
    state: str | None = None,
    source_role: str | None = None,
) -> dict[str, Any]:
    """Create a schema-shaped import record with explicit unknown metadata."""

    now = utc_now()
    normalized_state = state or initial_state(metadata)
    if normalized_state not in IMPORT_STATES:
        raise ValueError(f"Unknown import state: {normalized_state}")
    return {
        "schema_version": IMPORT_SCHEMA_VERSION,
        "kind": "neurobench_dataset_import",
        "import_id": str(import_id_value),
        "dataset_id": normalize_dataset_id(dataset_id),
        "source_mode": str(source_mode),
        "source_role": str(
            source_role
            or (
                "primary_video_candidate"
                if metadata.get("kind") == "video"
                else "neurev_json_attachment"
                if metadata.get("kind") == "neurev_json"
                else "label_attachment"
            )
        ),
        "is_primary_video": False,
        "original_name": str(original_name),
        "source_path": str(source_path),
        "destination_path": str(destination_path),
        "checksum": {
            "sha256": metadata.get("sha256"),
            "size_bytes": metadata.get("size_bytes"),
        },
        "state": normalized_state,
        "metadata": dict(metadata),
        "warnings": [str(item) for item in warnings],
        "generated_artifacts": {},
        "revision": 1,
        "created_at": now,
        "updated_at": now,
    }


def update_import_record(record: Mapping[str, Any], **updates: Any) -> dict[str, Any]:
    """Return an atomic-update-ready record while preserving immutable identity."""

    updated = dict(record)
    expected_revision = updates.pop("expected_revision", None)
    current_revision = int(record.get("revision") or 1)
    if expected_revision is not None and int(expected_revision) != current_revision:
        raise ValueError(f"Import revision changed (expected {int(expected_revision)}, found {current_revision}).")
    immutable = {key: record.get(key) for key in ("import_id", "dataset_id", "source_mode", "original_name", "source_path", "destination_path", "app_dir", "checksum", "created_at") if key in record}
    for key, value in updates.items():
        if key in immutable:
            raise ValueError(f"Import field '{key}' is immutable.")
        updated[key] = value
    updated.update(immutable)
    source_role = str(updated.get("source_role") or "")
    if source_role not in {"primary_video_candidate", "primary_video", "label_attachment", "neurev_json_attachment"}:
        raise ValueError(f"Unknown import source role: {source_role}")
    if source_role == "primary_video" and (updated.get("metadata") or {}).get("kind") != "video":
        raise ValueError("Only a video import can be promoted to primary_video.")
    updated["revision"] = current_revision + 1
    updated["updated_at"] = utc_now()
    if updated.get("state") not in IMPORT_STATES:
        raise ValueError(f"Unknown import state: {updated.get('state')}")
    return updated


def transition_import_record(
    record: Mapping[str, Any],
    state: str,
    *,
    expected_revision: int | None = None,
    **updates: Any,
) -> dict[str, Any]:
    """Apply one legal lifecycle transition with optional optimistic locking."""

    current = str(record.get("state") or "")
    current_label = current or "(missing)"
    if state not in IMPORT_STATES:
        raise ValueError(f"Unknown import state: {state}")
    if state not in IMPORT_STATE_TRANSITIONS.get(current, frozenset()):
        raise ValueError(f"Illegal import state transition: {current_label} -> {state}")
    return update_import_record(
        record,
        state=state,
        expected_revision=expected_revision,
        **updates,
    )


def verify_source_identity(path: str | Path, record: Mapping[str, Any]) -> dict[str, Any]:
    """Verify that a mutable registered path still matches the recorded source."""

    observed = checksum_file(path)
    expected = record.get("checksum") if isinstance(record.get("checksum"), Mapping) else {}
    expected_size = expected.get("size_bytes")
    expected_sha = expected.get("sha256")
    if expected_size is not None and int(expected_size) != int(observed["size_bytes"]):
        raise ValueError("Import source size changed after registration.")
    if expected_sha and str(expected_sha) != str(observed["sha256"]):
        raise ValueError("Import source checksum changed after registration.")
    return observed


def dataset_manifest_from_import(record: Mapping[str, Any], *, app_dir: str | Path) -> dict[str, Any]:
    """Create a truthful dataset manifest from an import record."""

    metadata = record.get("metadata") if isinstance(record.get("metadata"), Mapping) else {}
    if metadata.get("kind") != "video":
        raise ValueError("Only an explicitly promoted primary video can create a dataset manifest.")
    app = Path(app_dir)
    manifest: dict[str, Any] = {
        "schema_version": 1,
        "dataset_id": str(record.get("dataset_id") or "dataset"),
        "name": str(record.get("original_name") or record.get("dataset_id") or "dataset"),
        "source": {
            "mode": record.get("source_mode"),
            "import_id": record.get("import_id"),
        },
        "paths": {
           "app_dir": str(app),
            "review_data": str(app / "review_data.json"),
            "annotations": str(app / "annotations.json"),
            "architecture_runs": str(app / "architecture_runs.json"),
       },
   }
    manifest["paths"]["raw_video"] = str(record.get("destination_path") or record.get("source_path") or "")
    for key in ("modality", "indicator", "frame_rate_hz", "pixel_size_microns"):
        if metadata.get(key) is not None:
            manifest[key] = metadata[key]
    return manifest


def atomic_write_import_record(path: str | Path, record: Mapping[str, Any]) -> Path:
    """Write an import record atomically."""

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("x", encoding="utf-8") as handle:
            handle.write(json.dumps(dict(record), indent=2, sort_keys=True) + "\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)
    return target
